"""Convert attendees.xlsx into a structured attendees.json.

Walks all four worksheets in the source workbook, reads the hyperlink target from
the Profile column (rather than the literal cell text 'Profile'), and cross-links
named contacts so a single name lookup returns a merged record.

Run before deploying: the JSON is what ships, not the xlsx.
"""
import json
import os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
# Source xlsx lives in the Northell CRM folder. Override with ATTENDEES_XLSX env var.
DEFAULT_XLSX = os.path.expanduser(
    '~/Desktop/Brain/! Northell/CRM/attendees-programmatic-pioneers-2026-06-02.xlsx'
)
XLSX_PATH = os.environ.get('ATTENDEES_XLSX', DEFAULT_XLSX)
OUT_PATH = os.path.join(HERE, 'data', 'attendees.json')


def cell_link(cell):
    return cell.hyperlink.target if cell.hyperlink else None


def parse_attendees(ws):
    """Sheet columns: # | Name | Title | Company | Profile URL | Notes | Met? | Follow-up."""
    out = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        num = row[0].value
        name = row[1].value
        if not name:
            continue
        out.append({
            'id': int(num) if num is not None else None,
            'name': str(name).strip(),
            'title': (row[2].value or '').strip() if row[2].value else '',
            'company': (row[3].value or '').strip() if row[3].value else '',
            'profile_url': cell_link(row[4]) or '',
            'notes': (row[5].value or '').strip() if row[5].value else '',
            'met': bool(row[6].value) if row[6].value else False,
            'follow_up': (row[7].value or '').strip() if row[7].value else '',
            'tags': [],
        })
    return out


def parse_trade_body(ws):
    """Columns: Name | Title | Org | Org abbrev (meaning) | Priority | Open angle | Profile | Met? | Follow-up."""
    out = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        name = row[0].value
        if not name:
            continue
        out.append({
            'name': str(name).strip(),
            'title': (row[1].value or '').strip() if row[1].value else '',
            'org': (row[2].value or '').strip() if row[2].value else '',
            'org_meaning': (row[3].value or '').strip() if row[3].value else '',
            'priority': (row[4].value or '').strip() if row[4].value else '',
            'open_angle': (row[5].value or '').strip() if row[5].value else '',
            'profile_url': cell_link(row[6]) or '',
            'met': bool(row[7].value) if row[7].value else False,
            'follow_up': (row[8].value or '').strip() if row[8].value else '',
        })
    return out


def parse_agency(ws):
    """Columns: Name | Title | Agency | Holdco | Tier | Lens | Profile | Met? | Follow-up.

    Rows where only column A is populated are section bars (e.g. 'Dentsu (9)') - skip those.
    """
    out = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        name = row[0].value
        title = row[1].value
        agency = row[2].value
        if not name:
            continue
        if not (title or agency):
            # Section bar like 'Dentsu (9)' or 'IPG (0 - ABSENT)' - skip
            continue
        if str(name).strip() == '(none)' or str(name).strip() == '(none in room)':
            continue
        out.append({
            'name': str(name).strip(),
            'title': (title or '').strip() if title else '',
            'agency': (agency or '').strip() if agency else '',
            'holdco': (row[3].value or '').strip() if row[3].value else '',
            'tier': (row[4].value or '').strip() if row[4].value else '',
            'lens': (row[5].value or '').strip() if row[5].value else '',
            'profile_url': cell_link(row[6]) or '',
            'met': bool(row[7].value) if row[7].value else False,
            'follow_up': (row[8].value or '').strip() if row[8].value else '',
        })
    return out


def parse_better_collective(ws):
    """The Better Collective sheet uses a free-form section layout. Sections:
    COMPANY SNAPSHOT, HOUSE OF BRANDS, FANREACH, IN THE ROOM TODAY,
    WHY THIS IS A CAPCLEAR ACCOUNT, ACTION BOARD, OPEN ANGLE FOR TODAY.
    """
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        rows.append(row)

    company = {}
    house_of_brands = []
    fanreach = {}
    in_room_contacts = []
    why_capclear = []
    action_board = []
    open_angle = ''

    section = None
    for row in rows:
        a = row[0].value
        b = row[1].value if len(row) > 1 else None
        if a is None and b is None:
            continue
        a_str = (str(a).strip() if a else '')
        b_str = (str(b).strip() if b else '')

        # Detect section headers (single column, all caps)
        if a_str and b is None and a_str == a_str.upper() and len(a_str) > 4:
            up = a_str.upper()
            if 'COMPANY SNAPSHOT' in up:
                section = 'company'
            elif 'HOUSE OF BRANDS' in up:
                section = 'house'
            elif 'FANREACH' in up:
                section = 'fanreach'
            elif 'IN THE ROOM TODAY' in up:
                section = 'inroom'
            elif 'CAPCLEAR ACCOUNT' in up:
                section = 'why'
            elif 'ACTION BOARD' in up:
                section = 'action'
            elif 'OPEN ANGLE' in up:
                section = 'open_angle'
            elif 'BETTER COLLECTIVE - ACCOUNT BRIEF' in up:
                section = None
            continue

        # In-room contacts: skip column header row 'Name | Title | Lens ...'
        if section == 'inroom' and a_str == 'Name' and b_str == 'Title':
            continue

        if section == 'company' and a_str:
            company[a_str.lower().replace(' ', '_')] = b_str
        elif section == 'house' and a_str:
            house_of_brands.append({'brand': a_str, 'description': b_str})
        elif section == 'fanreach' and a_str:
            key = a_str.lower().replace(' ', '_').replace('/', '_').strip('_')
            fanreach[key] = b_str
        elif section == 'inroom' and a_str:
            in_room_contacts.append({
                'name': a_str,
                'title': b_str,
                'lens': (str(row[2].value).strip() if row[2].value else ''),
                'notes': (str(row[3].value).strip() if row[3].value else ''),
                'profile_url': cell_link(row[4]) if len(row) > 4 else '',
                'met': bool(row[5].value) if len(row) > 5 and row[5].value else False,
                'follow_up': (str(row[6].value).strip() if len(row) > 6 and row[6].value else ''),
            })
        elif section == 'why' and a_str:
            why_capclear.append({'reason': a_str, 'detail': b_str})
        elif section == 'action' and a_str:
            action_board.append({'when': a_str, 'action': b_str})
        elif section == 'open_angle' and a_str:
            open_angle = (open_angle + ' ' + a_str).strip() if open_angle else a_str

    return {
        'company': company,
        'house_of_brands': house_of_brands,
        'fanreach': fanreach,
        'in_room_contacts': in_room_contacts,
        'why_capclear_account': why_capclear,
        'action_board': action_board,
        'open_angle': open_angle,
    }


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)

    attendees = parse_attendees(wb['Attendees'])
    trade_body = parse_trade_body(wb['Trade Body Targets'])
    agency = parse_agency(wb['Agency Targets'])
    better_collective = parse_better_collective(wb['Better Collective'])

    # Cross-reference: tag attendees with trade-body / agency / better-collective context.
    name_to_attendee = {a['name'].lower(): a for a in attendees}
    for t in trade_body:
        a = name_to_attendee.get(t['name'].lower())
        if a:
            a['tags'].append('trade_body_target')
    for ag in agency:
        a = name_to_attendee.get(ag['name'].lower())
        if a:
            a['tags'].append('agency_target')
            if ag['tier']:
                a['tags'].append(ag['tier'].lower().replace(' ', '_'))
    for ic in better_collective['in_room_contacts']:
        a = name_to_attendee.get(ic['name'].lower())
        if a:
            a['tags'].append('better_collective_contact')

    out = {
        'metadata': {
            'conference': 'Programmatic Pioneers Summit',
            'date': '2026-06-02',
            'location': 'London',
            'organiser': 'WBR',
            'host': 'Trish Lynch',
            'captured_by': 'Andy Hosie, Head of Strategy, Northell',
            'total_attendees': len(attendees),
        },
        'attendees': attendees,
        'trade_body_targets': trade_body,
        'agency_targets': agency,
        'better_collective': better_collective,
    }

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, 'w') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    tier1 = sum(1 for a in agency if a['tier'] == 'TIER 1')
    print(f'Wrote {OUT_PATH}')
    print(f'  attendees: {len(attendees)}')
    print(f'  trade_body_targets: {len(trade_body)}')
    print(f'  agency_targets: {len(agency)} (Tier 1: {tier1})')
    print(f'  better_collective.in_room_contacts: {len(better_collective["in_room_contacts"])}')


if __name__ == '__main__':
    main()
